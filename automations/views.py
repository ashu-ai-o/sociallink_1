from django.shortcuts import render

# Create your views here.
"""
REST API Views for Automations with OpenRouter AI
Provides endpoints for testing and managing AI-enhanced automations
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.conf import settings
import asyncio
from asgiref.sync import async_to_sync

from .models import Automation, AutomationTrigger, Contact
from .serializers import (
    AutomationSerializer, 
    AutomationTriggerSerializer, 
    ContactSerializer
)
from .services.ai_service_async import (
    AIServiceOpenRouter,
    AIServiceOpenRouterSync
)
from .tasks import process_automation_trigger_async

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse, StreamingHttpResponse
from datetime import datetime, timedelta, timezone
from django.db.models import Sum


class AutomationViewSet(viewsets.ModelViewSet):
    """
    API endpoints for managing automations
    Includes AI enhancement testing
    """
    serializer_class = AutomationSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter automations by user's Instagram accounts"""
        return Automation.objects.filter(
            instagram_account__user=self.request.user
        ).select_related('instagram_account')
    
    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        """
        Toggle automation on/off
        POST /api/automations/{id}/toggle/
        """
        automation = self.get_object()
        automation.is_active = not automation.is_active
        automation.save()
        
        return Response({
            'id': str(automation.id),
            'name': automation.name,
            'is_active': automation.is_active,
            'message': f"Automation {'activated' if automation.is_active else 'deactivated'}"
        })
    
    @action(detail=True, methods=['post'])
    def test_trigger(self, request, pk=None):
        """
        Test automation trigger manually
        POST /api/automations/{id}/test_trigger/
        
        Body:
        {
            "comment_text": "I want this!",
            "instagram_username": "test_user",
            "instagram_user_id": "123456"
        }
        """
        automation = self.get_object()
        
        # Create test trigger
        trigger = AutomationTrigger.objects.create(
            automation=automation,
            instagram_user_id=request.data.get('instagram_user_id', '999999'),
            instagram_username=request.data.get('instagram_username', 'test_user'),
            comment_text=request.data.get('comment_text', 'Test comment'),
            status='pending'
        )
        
        # Queue for processing
        process_automation_trigger_async.delay(str(trigger.id))
        
        return Response({
            'success': True,
            'trigger_id': str(trigger.id),
            'message': 'Test trigger created and queued for processing',
            'automation': automation.name,
            'will_use_ai': automation.use_ai_enhancement
        })
    
    @action(detail=False, methods=['post'])
    def test_ai_enhancement(self, request):
        """
        Test AI message enhancement without creating automation
        POST /api/automations/test_ai_enhancement/
        
        Body:
        {
            "base_message": "Thanks for your comment!",
            "business_context": "We sell premium coffee",
            "user_comment": "This looks amazing!",
            "username": "coffee_lover"
        }
        """
        # Use synchronous version for REST API
        ai_service = AIServiceOpenRouterSync(
            api_key=settings.OPENROUTER_API_KEY,
            site_url=settings.OPENROUTER_SITE_URL
        )
        
        result = ai_service.enhance_dm_message(
            base_message=request.data.get('base_message', ''),
            business_context=request.data.get('business_context', ''),
            user_comment=request.data.get('user_comment', ''),
            username=request.data.get('username', 'user')
        )
        
        return Response({
            'success': result['success'],
            'original_message': result['original_message'],
            'enhanced_message': result['enhanced_message'],
            'model_used': result.get('model_name', 'N/A'),
            'provider': result.get('provider', 'openrouter'),
            'error': result.get('error')
        })
    
    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        """
        Get analytics for specific automation
        GET /api/automations/{id}/analytics/?days=30
        """
        automation = self.get_object()
        days = int(request.query_params.get('days', 30))
        
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Count, Q
        
        start_date = timezone.now() - timedelta(days=days)
        
        triggers = AutomationTrigger.objects.filter(
            automation=automation,
            created_at__gte=start_date
        )
        
        analytics = {
            'automation_id': str(automation.id),
            'automation_name': automation.name,
            'period_days': days,
            'total_triggers': triggers.count(),
            'triggers_by_status': {
                'sent': triggers.filter(status='sent').count(),
                'failed': triggers.filter(status='failed').count(),
                'skipped': triggers.filter(status='skipped').count(),
                'pending': triggers.filter(status='pending').count(),
            },
            'ai_enhanced_count': triggers.filter(was_ai_enhanced=True).count(),
            'success_rate': automation.total_dms_sent / automation.total_triggers * 100 
                           if automation.total_triggers > 0 else 0,
            'total_dms_sent': automation.total_dms_sent,
            'total_triggers': automation.total_triggers,
        }
        
        return Response(analytics)
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """
        Duplicate an automation
        POST /api/automations/{id}/duplicate/
        """
        automation = self.get_object()
        
        # Create copy
        new_automation = Automation.objects.create(
            instagram_account=automation.instagram_account,
            name=f"{automation.name} (Copy)",
            trigger_type=automation.trigger_type,
            trigger_keywords=automation.trigger_keywords,
            trigger_match_type=automation.trigger_match_type,
            target_posts=automation.target_posts,
            dm_message=automation.dm_message,
            dm_buttons=automation.dm_buttons,
            require_follow=automation.require_follow,
            follow_check_message=automation.follow_check_message,
            use_ai_enhancement=automation.use_ai_enhancement,
            ai_context=automation.ai_context,
            max_triggers_per_user=automation.max_triggers_per_user,
            cooldown_minutes=automation.cooldown_minutes,
            is_active=False,  # Start inactive
            priority=automation.priority,
        )
        
        serializer = self.get_serializer(new_automation)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AutomationTriggerViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoints for viewing automation triggers
    Read-only - triggers are created automatically
    """
    serializer_class = AutomationTriggerSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter triggers by user's automations"""
        return AutomationTrigger.objects.filter(
            automation__instagram_account__user=self.request.user
        ).select_related('automation').order_by('-created_at')
    


    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export triggers to CSV or Excel
        GET /api/triggers/export/?format=csv&automation_id=xxx
        """
        format_type = request.query_params.get('format', 'csv').lower()
        automation_id = request.query_params.get('automation_id')
        
        triggers = self.get_queryset()
        if automation_id:
            triggers = triggers.filter(automation_id=automation_id)
        
        fields = [
            ('automation.name', 'Automation'),
            ('instagram_username', 'Username'),
            ('comment_text', 'Comment'),
            ('status', 'Status'),
            ('dm_sent_at', 'DM Sent At'),
            ('was_ai_enhanced', 'AI Enhanced'),
            ('created_at', 'Triggered At'),
        ]
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'xlsx':
            filename = f'triggers_{timestamp}.xlsx'
            return export_to_excel(triggers, fields, filename, 'Triggers')
        else:
            filename = f'triggers_{timestamp}.csv'
            return export_to_csv(triggers, fields, filename)




class ContactViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoints for viewing contacts
    """
    serializer_class = ContactSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter contacts by user's Instagram accounts"""
        return Contact.objects.filter(
            instagram_account__user=self.request.user
        ).select_related('instagram_account')
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Search contacts by username
        GET /api/contacts/search/?q=username
        """
        query = request.query_params.get('q', '')
        contacts = self.get_queryset().filter(
            instagram_username__icontains=query
        )[:20]
        
        serializer = self.get_serializer(contacts, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export contacts to CSV/XLSX
        GET /api/contacts/export/?format=csv
        """
        format_type = request.query_params.get('format', 'csv')
        
        # This would generate CSV/XLSX file
        # For now, return JSON
        contacts = self.get_queryset()
        serializer = self.get_serializer(contacts, many=True)
        
        return Response({
            'format': format_type,
            'count': contacts.count(),
            'data': serializer.data,
            'message': 'Export functionality - implement with pandas or openpyxl'
        })
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export contacts to CSV or Excel
        GET /api/contacts/export/?format=csv
        GET /api/contacts/export/?format=xlsx
        """
        format_type = request.query_params.get('format', 'csv').lower()
        contacts = self.get_queryset()
        
        # Define fields to export
        fields = [
            ('instagram_username', 'Username'),
            ('full_name', 'Full Name'),
            ('instagram_user_id', 'Instagram ID'),
            ('total_interactions', 'Total Interactions'),
            ('total_dms_received', 'DMs Received'),
            ('is_follower', 'Is Follower'),
            ('first_interaction', 'First Interaction'),
            ('last_interaction', 'Last Interaction'),
            ('tags', 'Tags'),
        ]
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'xlsx':
            filename = f'contacts_{timestamp}.xlsx'
            return export_to_excel(contacts, fields, filename, 'Contacts')
        else:
            filename = f'contacts_{timestamp}.csv'
            return export_to_csv(contacts, fields, filename)




class AIProviderViewSet(viewsets.ViewSet):
    """
    API endpoints for testing AI providers
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def status(self, request):
        """
        Get status of all AI providers/models
        GET /api/ai-providers/status/
        """
        from .services.ai_service_async import AIServiceOpenRouter
        
        models_info = AIServiceOpenRouter.MODELS
        
        return Response({
            'provider': 'OpenRouter',
            'api_url': settings.OPENROUTER_API_URL,
            'models_configured': len(models_info),
            'models': [
                {
                    'id': m['id'],
                    'name': m['name'],
                    'cost': m['cost'],
                    'speed': m['speed'],
                    'use_case': m['use_case']
                }
                for m in models_info
            ]
        })
    
    @action(detail=False, methods=['post'])
    def test(self, request):
        """
        Test AI provider with custom message
        POST /api/ai-providers/test/
        
        Body:
        {
            "prompt": "Your test prompt",
            "model": "anthropic/claude-3.5-sonnet"  // optional
        }
        """
        ai_service = AIServiceOpenRouterSync(
            api_key=settings.OPENROUTER_API_KEY,
            site_url=settings.OPENROUTER_SITE_URL
        )
        
        prompt = request.data.get('prompt', 'Hello, how are you?')
        model = request.data.get('model')
        
        # Test generation
        import time
        start_time = time.time()
        
        result = ai_service._generate(
            prompt=prompt,
            model=model or ai_service.MODELS[0]['id']
        )
        
        elapsed = time.time() - start_time
        
        return Response({
            'success': result['success'],
            'response': result.get('text', 'N/A'),
            'model': model or ai_service.MODELS[0]['id'],
            'response_time': f"{elapsed:.2f}s",
            'error': result.get('error')
        })


class AnalyticsViewSet(viewsets.ViewSet):
    """
    API endpoints for analytics and dashboard stats
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """
        Get dashboard statistics
        GET /api/analytics/dashboard/?period=30d
        """
        from django.db.models import Sum, Count, Q
        from django.utils import timezone
        from datetime import timedelta
        
        period = request.query_params.get('period', '30d')
        days = int(period.replace('d', ''))
        
        start_date = timezone.now() - timedelta(days=days)
        
        automations = Automation.objects.filter(
            instagram_account__user=request.user
        )
        
        triggers = AutomationTrigger.objects.filter(
            automation__instagram_account__user=request.user,
            created_at__gte=start_date
        )
        
        stats = {
            'period': period,
            'total_automations': automations.count(),
            'active_automations': automations.filter(is_active=True).count(),
            'total_dms_sent': automations.aggregate(
                total=Sum('total_dms_sent')
            )['total'] or 0,
            'total_triggers': triggers.count(),
            'today_triggers': triggers.filter(
                created_at__date=timezone.now().date()
            ).count(),
            'success_rate': (
                triggers.filter(status='sent').count() / triggers.count() * 100
                if triggers.count() > 0 else 0
            ),
            'ai_enhanced_rate': (
                triggers.filter(was_ai_enhanced=True).count() / triggers.count() * 100
                if triggers.count() > 0 else 0
            ),
            'daily_breakdown': self._get_daily_breakdown(triggers, days)
        }
        
        return Response(stats)
    
    def _get_daily_breakdown(self, triggers, days):
        """Get daily breakdown of triggers"""
        from django.utils import timezone
        from datetime import timedelta
        
        breakdown = []
        for i in range(days):
            date = timezone.now().date() - timedelta(days=days-i-1)
            day_triggers = triggers.filter(created_at__date=date)
            
            breakdown.append({
                'date': date.isoformat(),
                'triggers': day_triggers.count(),
                'dms_sent': day_triggers.filter(status='sent').count(),
                'ai_enhanced': day_triggers.filter(was_ai_enhanced=True).count()
            })
        
        return breakdown
    
    @action(detail=False, methods=['get'])
    def automations(self, request):
        """
        Get performance metrics for all automations
        GET /api/analytics/automations/
        """
        automations = Automation.objects.filter(
            instagram_account__user=request.user
        )
        
        performance = []
        for automation in automations:
            performance.append({
                'id': str(automation.id),
                'name': automation.name,
                'is_active': automation.is_active,
                'total_triggers': automation.total_triggers,
                'total_dms_sent': automation.total_dms_sent,
                'success_rate': (
                    automation.total_dms_sent / automation.total_triggers * 100
                    if automation.total_triggers > 0 else 0
                ),
                'conversions': automation.total_dms_sent  # Placeholder
            })
        
        return Response({
            'count': len(performance),
            'top_automations': sorted(
                performance, 
                key=lambda x: x['total_dms_sent'], 
                reverse=True
            )[:10]
        })
    



    @action(detail=False, methods=['get'])
    def export_analytics(self, request):
        """
        Export analytics to Excel with multiple sheets
        GET /api/analytics/dashboard/export_analytics/?period=30d
        """
        from automations.models import Automation, AutomationTrigger
        
        period = request.query_params.get('period', '30d')
        days = int(period.replace('d', ''))
        start_date = timezone.now() - timedelta(days=days)
        
        workbook = Workbook()
        
        # Sheet 1: Overview
        ws_overview = workbook.active
        ws_overview.title = 'Overview'
        
        automations = Automation.objects.filter(
            instagram_account__user=request.user
        )
        
        overview_data = [
            ['Metric', 'Value'],
            ['Total Automations', automations.count()],
            ['Active Automations', automations.filter(is_active=True).count()],
            ['Total DMs Sent', automations.aggregate(Sum('total_dms_sent'))['total_dms_sent__sum'] or 0],
            ['Total Triggers', automations.aggregate(Sum('total_triggers'))['total_triggers__sum'] or 0],
            ['Period', f'{days} days'],
            ['Report Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row_num, row_data in enumerate(overview_data, 1):
            for col_num, value in enumerate(row_data, 1):
                ws_overview.cell(row=row_num, column=col_num, value=value)
        
        # Sheet 2: Automation Performance
        ws_automations = workbook.create_sheet('Automations')
        ws_automations.append(['Name', 'Status', 'Triggers', 'DMs Sent', 'Success Rate'])
        
        for automation in automations:
            success_rate = (automation.total_dms_sent / automation.total_triggers * 100) if automation.total_triggers > 0 else 0
            ws_automations.append([
                automation.name,
                'Active' if automation.is_active else 'Inactive',
                automation.total_triggers,
                automation.total_dms_sent,
                f'{success_rate:.1f}%'
            ])
        
        # Sheet 3: Daily Breakdown
        ws_daily = workbook.create_sheet('Daily Breakdown')
        ws_daily.append(['Date', 'Triggers', 'DMs Sent', 'AI Enhanced'])
        
        triggers = AutomationTrigger.objects.filter(
            automation__instagram_account__user=request.user,
            created_at__gte=start_date
        )
        
        for i in range(days):
            date = (timezone.now() - timedelta(days=days-i-1)).date()
            day_triggers = triggers.filter(created_at__date=date)
            
            ws_daily.append([
                date.strftime('%Y-%m-%d'),
                day_triggers.count(),
                day_triggers.filter(status='sent').count(),
                day_triggers.filter(was_ai_enhanced=True).count()
            ])
        
        # Style headers
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
        
        # Save
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'analytics_report_{timestamp}.xlsx'
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    


















# ============================================================================
# CSV EXPORT HELPERS
# ============================================================================

class Echo:
    """Helper for streaming CSV responses"""
    def write(self, value):
        return value


def export_to_csv(queryset, fields, filename):
    """
    Generic CSV export function
    
    Args:
        queryset: Django queryset to export
        fields: List of (field_name, display_name) tuples
        filename: Output filename
    """
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer)
    
    def generate_rows():
        # Header row
        yield writer.writerow([field[1] for field in fields])
        
        # Data rows
        for obj in queryset:
            row = []
            for field_name, _ in fields:
                # Handle nested fields (e.g., 'automation.name')
                if '.' in field_name:
                    parts = field_name.split('.')
                    value = obj
                    for part in parts:
                        value = getattr(value, part, '')
                else:
                    value = getattr(obj, field_name, '')
                
                # Format dates
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                
                row.append(str(value))
            yield writer.writerow(row)
    
    response = StreamingHttpResponse(
        generate_rows(),
        content_type='text/csv'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# EXCEL EXPORT HELPERS
# ============================================================================

def export_to_excel(queryset, fields, filename, sheet_name='Data'):
    """
    Generic Excel export function with styling
    
    Args:
        queryset: Django queryset to export
        fields: List of (field_name, display_name) tuples
        filename: Output filename
        sheet_name: Excel sheet name
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    
    # Header styling
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Write headers
    for col_num, (_, display_name) in enumerate(fields, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = display_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (field_name, _) in enumerate(fields, 1):
            # Handle nested fields
            if '.' in field_name:
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    value = getattr(value, part, '')
            else:
                value = getattr(obj, field_name, '')
            
            # Format dates
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            
            worksheet.cell(row=row_num, column=col_num, value=str(value))
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
